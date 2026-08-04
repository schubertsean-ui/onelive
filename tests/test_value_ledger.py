"""Tests for tools/value_ledger.py — the Agent Value Ledger.

Hermetic: every case runs on a tmp_path workbook with caller-supplied dates
(the tool never reads the wall clock); no network, no repo mutation. The one
repo-facing case reads the COMMITTED workbook+mirror pair read-only and proves
they agree — the property (contents consistent), not that a tool ran
(§9.6 "a gate that cannot fail proves nothing"). Every refusal branch has its
own red case: the ledger fails closed on bad numbers, bad dates, empty
estimate basis, tampered headers, missing sheets, and stale derived surfaces.
"""
import pathlib
from decimal import Decimal

import openpyxl
import pytest

import tools.value_ledger as vl

REPO = pathlib.Path(__file__).resolve().parent.parent


@pytest.fixture
def book(tmp_path):
    path = tmp_path / "ledger.xlsx"
    assert vl.main(["--path", str(path), "init"]) == 0
    return path


def _log(path, **over):
    args = {"date": "2026-08-04", "session": "sync-run-001", "task": "sync client calendar",
            "category": "sync", "hours": "6", "basis": "time a comparable manual sync takes"}
    args.update(over)
    return vl.main(["--path", str(path), "log"] + [
        item for k, v in args.items() for item in (f"--{k}", v)])


def test_init_creates_schema_and_mirror(book):
    wb = openpyxl.load_workbook(book)
    assert set(wb.sheetnames) == {vl.LEDGER_SHEET, vl.WEEKLY_SHEET, vl.CONFIG_SHEET}
    assert [c.value for c in next(wb[vl.LEDGER_SHEET].iter_rows(max_row=1))] == vl.LEDGER_HEADERS
    assert vl.read_rate(wb) == vl.DEFAULT_RATE
    assert book.with_suffix(".csv").read_text().startswith(",".join(vl.LEDGER_HEADERS))


def test_init_refuses_overwrite(book):
    assert vl.main(["--path", str(book), "init"]) == 1


def test_log_appends_value_and_regenerates_derived(book):
    assert _log(book) == 0
    wb = openpyxl.load_workbook(book)
    entries = vl.read_entries(wb)
    assert len(entries) == 1
    assert entries[0]["value_usd"] == Decimal("6") * vl.DEFAULT_RATE
    assert entries[0]["iso_week"] == "2026-W32"
    weekly = [list(r) for r in wb[vl.WEEKLY_SHEET].iter_rows(values_only=True)]
    assert weekly[0] == vl.WEEKLY_HEADERS
    assert weekly[1][:3] == ["2026-W32", "2026-08-03", 1]
    assert vl.main(["--path", str(book), "verify"]) == 0


def test_weekly_rollup_and_cumulative_across_weeks(book):
    assert _log(book, date="2026-08-04", hours="2") == 0
    assert _log(book, date="2026-08-05", hours="3") == 0
    assert _log(book, date="2026-08-12", hours="1.5") == 0  # next ISO week
    rollup = vl.weekly_rollup(vl.read_entries(openpyxl.load_workbook(book)))
    assert [w["iso_week"] for w in rollup] == ["2026-W32", "2026-W33"]
    assert rollup[0]["hours_saved"] == Decimal("5")
    assert rollup[1]["cumulative_value_usd"] == Decimal("6.5") * vl.DEFAULT_RATE


def test_founder_edited_rate_applies_to_future_rows_only(book):
    assert _log(book, date="2026-08-04", hours="1") == 0
    wb = openpyxl.load_workbook(book)
    for row in wb[vl.CONFIG_SHEET].iter_rows(max_col=2):
        if row[0].value == "hourly_rate_usd":
            row[1].value = 200  # the founder edits the shared workbook in Excel
    wb.save(book)
    assert _log(book, date="2026-08-05", hours="1") == 0
    entries = vl.read_entries(openpyxl.load_workbook(book))
    assert [e["value_usd"] for e in entries] == [Decimal("150.00"), Decimal("200.00")]


@pytest.mark.parametrize("bad", ["nan", "inf", "-1", "abc", ""])
def test_refuses_nonfinite_negative_or_nonnumeric_hours(book, bad):
    assert _log(book, hours=bad) == 1


@pytest.mark.parametrize("bad", ["08/04/2026", "2026-13-01", "2026-8-4", "tomorrow"])
def test_refuses_malformed_dates(book, bad):
    assert _log(book, date=bad) == 1


def test_refuses_empty_estimate_basis_and_task(book):
    assert _log(book, basis="  ") == 1
    assert _log(book, task="") == 1


def test_refuses_missing_workbook(tmp_path):
    assert _log(tmp_path / "absent.xlsx") == 1
    assert vl.main(["--path", str(tmp_path / "absent.xlsx"), "verify"]) == 1


def test_refuses_tampered_header_and_missing_sheet(book):
    wb = openpyxl.load_workbook(book)
    wb[vl.LEDGER_SHEET]["A1"] = "renamed"
    wb.save(book)
    assert _log(book) == 1
    wb = openpyxl.load_workbook(book)
    wb[vl.LEDGER_SHEET]["A1"] = "date"  # restore header, then remove a sheet
    del wb[vl.CONFIG_SHEET]
    wb.save(book)
    assert _log(book) == 1


def test_refuses_nonpositive_config_rate(book):
    wb = openpyxl.load_workbook(book)
    for row in wb[vl.CONFIG_SHEET].iter_rows(max_col=2):
        if row[0].value == "hourly_rate_usd":
            row[1].value = 0
    wb.save(book)
    assert _log(book) == 1


def test_refuses_tampered_ledger_value(book):
    assert _log(book, hours="2") == 0
    wb = openpyxl.load_workbook(book)
    wb[vl.LEDGER_SHEET].cell(row=2, column=vl.LEDGER_HEADERS.index("value_usd") + 1,
                             value=999999.0)
    wb.save(book)
    assert vl.main(["--path", str(book), "verify"]) == 1  # value != hours x rate


def test_verify_refuses_stale_mirror(book):
    assert _log(book) == 0
    book.with_suffix(".csv").write_text("tampered\n")
    assert vl.main(["--path", str(book), "verify"]) == 1


def test_report_prints_weekly_and_cumulative(book, capsys):
    assert _log(book, date="2026-08-04", hours="2") == 0
    assert _log(book, date="2026-07-30", hours="4") == 0  # prior ISO week
    assert vl.main(["--path", str(book), "report", "--as-of", "2026-08-04"]) == 0
    out = capsys.readouterr().out
    assert "This week (2026-W32): 1 task(s), ~2.00 hours saved ≈ $300.00" in out
    assert "Last week (2026-W31): 1 task(s), ~4.00 hours saved ≈ $600.00" in out
    assert "All time: 2 task(s)" in out
    assert "estimates" in out  # the report never claims actual/invoiced dollars


def test_committed_demo_ledger_is_consistent():
    """The demo client workbook committed in docs/strategy/examples/ must
    always verify (workbook, weekly sheet, and CSV mirror agree)."""
    assert vl.main(["--path", str(vl.DEFAULT_XLSX), "verify"]) == 0


def test_committed_demo_row_is_unmistakably_demo():
    """Two lessons pinned at once. Evaluator catch on PR #159 r1: a committed
    ledger row silently carried a stale identifier after a renumbering — so
    the committed artifact's row identity is PINNED (whoever changes it must
    re-sweep this pin deliberately, in the same change). Founder catch on the
    same PR: this ledger belongs to the Owned Agent's CLIENTS, not the repo's
    agent org — so the only committed row must be unmistakably illustrative
    (claim-ledger discipline: examples never read as live client work)."""
    wb = vl._load_workbook(vl.DEFAULT_XLSX)
    entries = vl.read_entries(wb)
    assert len(entries) == 1
    first = entries[0]
    assert first["session"] == "demo-seed"
    assert first["task"].startswith("DEMO (illustrative):")
    assert "DEMO row" in first["estimate_basis"]
