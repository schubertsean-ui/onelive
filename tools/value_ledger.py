#!/usr/bin/env python3
"""value_ledger — the Agent Value Ledger (founder-directed 2026-08-04).

SUMMARY: logs every completed agent task — hours saved and $ value — into a
shared EXCEL workbook (`docs/metrics/AGENT_VALUE_LEDGER.xlsx`), maintains a
"Weekly ROI" sheet, and prints the founder-facing weekly "you saved $" report.
The workbook is the founder-directed shared source of truth ("Make the agent
write to Excel, not its own markdown"): the founder opens and edits it in
Excel (the hourly rate lives in its Config sheet); the agent appends rows with
this tool. Because a binary xlsx is opaque to git diffs and to the non-Claude
evaluator, every write regenerates a deterministic CSV mirror
(`AGENT_VALUE_LEDGER.csv`) of the Ledger sheet — the xlsx stays canonical, the
CSV is the audit mirror, and `verify` (plus tests) refuses when they disagree.

Honesty physics, not policy:
  - Every hours-saved figure is an ESTIMATE and each row must carry its
    estimate basis — a row with no stated basis is REFUSED (an unexplained
    number is a fabrication vector, `[S3:fabricated-qualitative-copy]`).
  - $ value = hours_saved x rate, computed in Decimal and quantized to cents;
    NaN/Infinity/negative hours or rate REFUSE (`[S3:nonfinite-numeric-accepted]`).
  - The rate each row used is frozen into the row, so a later Config edit
    never silently rewrites history.
  - A workbook with missing sheets or tampered headers REFUSES loudly with
    the defect named — nothing silently recreates or skips
    (`[S3:swallowed-corrupt-data]`).
  - Deterministic: dates are caller-supplied (`--date`/`--as-of`); this
    module never reads the wall clock (kpi_report convention).

This ledger is bookkeeping (docs/metrics/): no gate reads it, it publishes
nothing, and it touches no product data path. Sending the weekly report
anywhere (email, group chat, scheduled runner) is a delivery CHANNEL —
new-service territory, founder-crucial by charter — and is deliberately not
in this tool; `report` only prints.

Usage:
  python tools/value_ledger.py init [--path P]
  python tools/value_ledger.py log --date YYYY-MM-DD --session S --task T \
      --category C --hours H --basis "how the estimate was made" [--path P]
  python tools/value_ledger.py report --as-of YYYY-MM-DD [--path P]
  python tools/value_ledger.py verify [--path P]

Exit codes (tools/README.md convention):
  0 = ok;  1 = refusal/defect (message names it);  2 = usage error.
"""
from __future__ import annotations

import argparse
import csv
import datetime as dt
import io
import pathlib
import re
import sys
from decimal import Decimal, InvalidOperation

REPO = pathlib.Path(__file__).resolve().parent.parent
DEFAULT_XLSX = REPO / "docs" / "metrics" / "AGENT_VALUE_LEDGER.xlsx"

LEDGER_SHEET = "Ledger"
WEEKLY_SHEET = "Weekly ROI"
CONFIG_SHEET = "Config"

LEDGER_HEADERS = [
    "date", "iso_week", "session", "task", "category",
    "hours_saved", "estimate_basis", "rate_usd_per_hour", "value_usd",
]
WEEKLY_HEADERS = [
    "iso_week", "week_start", "tasks_logged", "hours_saved",
    "value_usd", "cumulative_value_usd",
]

# Shipped default rate — a labeled PLACEHOLDER, not a measured fact. The
# founder sets the real blended rate directly in the workbook's Config sheet
# (that is the shared-source-of-truth flow the directive asked for).
DEFAULT_RATE = Decimal("150")
RATE_NOTE = (
    "PLACEHOLDER default - founder: set your real blended hourly rate in the "
    "cell above. $ values everywhere in this workbook are ESTIMATES "
    "(hours_saved x rate), never invoiced/actual dollars. Each Ledger row "
    "freezes the rate it was logged under, so changing this cell only "
    "affects future rows."
)

CENTS = Decimal("0.01")
DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")


class LedgerError(Exception):
    """A named refusal — the ledger fails closed, loudly."""


def _openpyxl():
    try:
        import openpyxl
    except ImportError as exc:  # loud, actionable — never a silent no-op
        raise LedgerError(
            "openpyxl is not installed (pinned in tools/requirements.txt): "
            f"pip install -r tools/requirements.txt ({exc})"
        ) from exc
    return openpyxl


def parse_date(text: str) -> dt.date:
    if not DATE_RE.match(text or ""):
        raise LedgerError(f"date must be YYYY-MM-DD, got {text!r}")
    try:
        return dt.date.fromisoformat(text)
    except ValueError as exc:
        raise LedgerError(f"invalid calendar date {text!r}: {exc}") from exc


def parse_nonneg_decimal(text: str, label: str) -> Decimal:
    """One validator for every numeric input: finite, non-negative Decimal."""
    try:
        value = Decimal(str(text).strip())
    except (InvalidOperation, ValueError) as exc:
        raise LedgerError(f"{label} is not a number: {text!r}") from exc
    if not value.is_finite():
        raise LedgerError(f"{label} must be finite, got {text!r}")
    if value < 0:
        raise LedgerError(f"{label} must be >= 0, got {text!r}")
    return value


def iso_week_of(day: dt.date) -> tuple[str, dt.date]:
    year, week, weekday = day.isocalendar()
    return f"{year}-W{week:02d}", day - dt.timedelta(days=weekday - 1)


def _require(condition: bool, defect: str) -> None:
    if not condition:
        raise LedgerError(defect)


# --- workbook schema ---------------------------------------------------------
def _load_workbook(path: pathlib.Path):
    _require(path.exists(), f"workbook not found: {path} (run `init` to create it)")
    wb = _openpyxl().load_workbook(path)
    for sheet in (LEDGER_SHEET, WEEKLY_SHEET, CONFIG_SHEET):
        _require(sheet in wb.sheetnames,
                 f"workbook {path} is missing sheet {sheet!r} — refusing "
                 "(fix or re-init deliberately; nothing is recreated silently)")
    header = [c.value for c in next(wb[LEDGER_SHEET].iter_rows(max_row=1))]
    _require(header == LEDGER_HEADERS,
             f"Ledger header mismatch: expected {LEDGER_HEADERS}, found {header}")
    return wb


def read_rate(wb) -> Decimal:
    ws = wb[CONFIG_SHEET]
    for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
        if row and row[0] == "hourly_rate_usd":
            rate = parse_nonneg_decimal(str(row[1]), "hourly_rate_usd (Config)")
            _require(rate > 0, f"hourly_rate_usd (Config) must be > 0, got {row[1]!r}")
            return rate
    raise LedgerError("Config sheet has no 'hourly_rate_usd' row — refusing")


def read_entries(wb) -> list[dict]:
    """Ledger rows as dicts of validated values — corrupt cells refuse loudly."""
    entries = []
    for i, row in enumerate(wb[LEDGER_SHEET].iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in row):
            continue
        rec = dict(zip(LEDGER_HEADERS, row))
        where = f"Ledger row {i}"
        for key in ("date", "session", "task", "category", "estimate_basis"):
            _require(bool(str(rec.get(key) or "").strip()), f"{where}: empty {key!r}")
        day = parse_date(str(rec["date"]))
        week, _ = iso_week_of(day)
        _require(rec["iso_week"] == week,
                 f"{where}: iso_week {rec['iso_week']!r} does not match date (expected {week})")
        rec["hours_saved"] = parse_nonneg_decimal(str(rec["hours_saved"]), f"{where}: hours_saved")
        rec["rate_usd_per_hour"] = parse_nonneg_decimal(
            str(rec["rate_usd_per_hour"]), f"{where}: rate_usd_per_hour")
        rec["value_usd"] = parse_nonneg_decimal(str(rec["value_usd"]), f"{where}: value_usd")
        expected = (rec["hours_saved"] * rec["rate_usd_per_hour"]).quantize(CENTS)
        _require(rec["value_usd"].quantize(CENTS) == expected,
                 f"{where}: value_usd {rec['value_usd']} != hours x rate ({expected})")
        entries.append(rec)
    return entries


def weekly_rollup(entries: list[dict]) -> list[dict]:
    weeks: dict[str, dict] = {}
    for rec in entries:
        week, start = iso_week_of(parse_date(str(rec["date"])))
        agg = weeks.setdefault(week, {
            "iso_week": week, "week_start": start.isoformat(),
            "tasks_logged": 0, "hours_saved": Decimal("0"), "value_usd": Decimal("0"),
        })
        agg["tasks_logged"] += 1
        agg["hours_saved"] += rec["hours_saved"]
        agg["value_usd"] += rec["value_usd"]
    out, cumulative = [], Decimal("0")
    for week in sorted(weeks):
        agg = weeks[week]
        cumulative += agg["value_usd"]
        agg["cumulative_value_usd"] = cumulative
        out.append(agg)
    return out


def mirror_csv_text(entries: list[dict]) -> str:
    """Deterministic CSV of the Ledger sheet — exact decimal strings."""
    buf = io.StringIO()
    writer = csv.writer(buf, lineterminator="\n")
    writer.writerow(LEDGER_HEADERS)
    for rec in entries:
        writer.writerow([
            rec["date"], rec["iso_week"], rec["session"], rec["task"], rec["category"],
            str(rec["hours_saved"].quantize(CENTS)),
            rec["estimate_basis"],
            str(rec["rate_usd_per_hour"].quantize(CENTS)),
            str(rec["value_usd"].quantize(CENTS)),
        ])
    return buf.getvalue()


def _rewrite_derived(wb, entries: list[dict], path: pathlib.Path) -> None:
    """Regenerate the Weekly ROI sheet + CSV mirror from the Ledger rows,
    then save. Derived surfaces are always rewritten wholesale — they are
    projections of the Ledger, never independent state."""
    ws = wb[WEEKLY_SHEET]
    ws.delete_rows(1, ws.max_row)
    ws.append(WEEKLY_HEADERS)
    for agg in weekly_rollup(entries):
        ws.append([
            agg["iso_week"], agg["week_start"], agg["tasks_logged"],
            float(agg["hours_saved"].quantize(CENTS)),
            float(agg["value_usd"].quantize(CENTS)),
            float(agg["cumulative_value_usd"].quantize(CENTS)),
        ])
    wb.save(path)
    path.with_suffix(".csv").write_text(mirror_csv_text(entries), encoding="utf-8")


# --- commands ----------------------------------------------------------------
def cmd_init(path: pathlib.Path) -> int:
    _require(not path.exists(), f"{path} already exists — refusing to overwrite")
    openpyxl = _openpyxl()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = LEDGER_SHEET
    ws.append(LEDGER_HEADERS)
    wb.create_sheet(WEEKLY_SHEET).append(WEEKLY_HEADERS)
    config = wb.create_sheet(CONFIG_SHEET)
    config.append(["hourly_rate_usd", float(DEFAULT_RATE)])
    config.append(["rate_note", RATE_NOTE])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    path.with_suffix(".csv").write_text(mirror_csv_text([]), encoding="utf-8")
    print(f"initialized {path} (+ CSV mirror); Config hourly_rate_usd={DEFAULT_RATE} (placeholder)")
    return 0


def cmd_log(path: pathlib.Path, date: str, session: str, task: str,
            category: str, hours: str, basis: str) -> int:
    day = parse_date(date)
    for label, text in (("--session", session), ("--task", task),
                        ("--category", category), ("--basis", basis)):
        _require(bool(text and text.strip()),
                 f"{label} must be non-empty — every entry states what it was and "
                 "(for --basis) how its hours-saved estimate was made")
    hours_dec = parse_nonneg_decimal(hours, "--hours").quantize(CENTS)
    wb = _load_workbook(path)
    read_entries(wb)  # refuse to append to a corrupt ledger
    rate = read_rate(wb)
    value = (hours_dec * rate).quantize(CENTS)
    week, _ = iso_week_of(day)
    wb[LEDGER_SHEET].append([
        day.isoformat(), week, session.strip(), task.strip(), category.strip(),
        float(hours_dec), basis.strip(), float(rate.quantize(CENTS)), float(value),
    ])
    _rewrite_derived(wb, read_entries(wb), path)
    print(f"logged: {day.isoformat()} [{category.strip()}] {task.strip()} — "
          f"{hours_dec}h saved x ${rate}/h = ${value} (estimate; basis: {basis.strip()})")
    return 0


def cmd_report(path: pathlib.Path, as_of: str) -> int:
    day = parse_date(as_of)
    entries = read_entries(_load_workbook(path))
    weeks = {w["iso_week"]: w for w in weekly_rollup(entries)}
    this_week, this_start = iso_week_of(day)
    last_week, _ = iso_week_of(this_start - dt.timedelta(days=1))
    total_hours = sum((e["hours_saved"] for e in entries), Decimal("0"))
    total_value = sum((e["value_usd"] for e in entries), Decimal("0"))

    def line(week: str, label: str) -> str:
        agg = weeks.get(week)
        if not agg:
            return f"{label} ({week}): no tasks logged."
        return (f"{label} ({week}): {agg['tasks_logged']} task(s), "
                f"~{agg['hours_saved'].quantize(CENTS)} hours saved "
                f"≈ ${agg['value_usd'].quantize(CENTS)}")

    print(f"AGENT VALUE — weekly ROI report (as of {day.isoformat()})")
    print(line(this_week, "This week"))
    print(line(last_week, "Last week"))
    print(f"All time: {len(entries)} task(s), ~{total_hours.quantize(CENTS)} hours saved "
          f"≈ ${total_value.quantize(CENTS)}")
    print("All figures are estimates (hours saved x the hourly rate each entry was "
          f"logged under; source of truth: {path.name}, founder-editable rate in Config).")
    return 0


def cmd_verify(path: pathlib.Path) -> int:
    wb = _load_workbook(path)
    entries = read_entries(wb)
    read_rate(wb)
    csv_path = path.with_suffix(".csv")
    _require(csv_path.exists(), f"CSV mirror missing: {csv_path}")
    expected = mirror_csv_text(entries)
    actual = csv_path.read_text(encoding="utf-8")
    _require(actual == expected,
             f"CSV mirror {csv_path.name} does not match the workbook's Ledger "
             "sheet — regenerate via any `log`, or investigate which was edited")
    ws = wb[WEEKLY_SHEET]
    got = [list(r) for r in ws.iter_rows(values_only=True)
           if any(v is not None for v in r)]
    want = [WEEKLY_HEADERS] + [
        [a["iso_week"], a["week_start"], a["tasks_logged"],
         float(a["hours_saved"].quantize(CENTS)), float(a["value_usd"].quantize(CENTS)),
         float(a["cumulative_value_usd"].quantize(CENTS))]
        for a in weekly_rollup(entries)]
    _require(got == want, f"'{WEEKLY_SHEET}' sheet is stale vs the Ledger rows — "
             "regenerate via any `log`, or investigate which was edited")
    print(f"verify OK: {len(entries)} entries; workbook, weekly sheet, and CSV mirror agree")
    return 0


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("--path", type=pathlib.Path, default=DEFAULT_XLSX)
    sub = parser.add_subparsers(dest="command", required=True)
    sub.add_parser("init")
    p_log = sub.add_parser("log")
    for flag in ("--date", "--session", "--task", "--category", "--hours", "--basis"):
        p_log.add_argument(flag, required=True)
    p_report = sub.add_parser("report")
    p_report.add_argument("--as-of", required=True)
    sub.add_parser("verify")
    args = parser.parse_args(argv)
    try:
        if args.command == "init":
            return cmd_init(args.path)
        if args.command == "log":
            return cmd_log(args.path, args.date, args.session, args.task,
                           args.category, args.hours, args.basis)
        if args.command == "report":
            return cmd_report(args.path, args.as_of)
        return cmd_verify(args.path)
    except LedgerError as exc:
        print(f"value_ledger REFUSED: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    sys.exit(main())
